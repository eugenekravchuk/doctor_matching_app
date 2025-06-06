import customtkinter as ctk
from tkinter import filedialog, messagebox
from algo_flow import generate_monthly_schedule_from_csv, generate_preference_schedule_from_csv, change_weekly_schedule
import os
import sys
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from collections import defaultdict
import threading
import time
import traceback

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

class ScheduleApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Schedule Management App")
        self.geometry("600x650")

        self.input_csv_one = None
        self.input_csv_two = None

        self.input_json = resource_path("data/rooms_locations_updated.json")
        self.weekly_schedule_file = None
        self.deleted_shifts_file = None

        self.main_frame = ctk.CTkFrame(self)
        self.main_frame.pack(pady=20, padx=20, fill="both", expand=True)

        ctk.CTkLabel(self.main_frame, text="Schedule Management", font=("Helvetica", 20, "bold")).pack(pady=10)

        monthly_frame = ctk.CTkFrame(self.main_frame)
        monthly_frame.pack(pady=10, padx=10, fill="x")
        ctk.CTkLabel(monthly_frame, text="Generate Monthly Schedule", font=("Helvetica", 14)).pack(pady=5)

        self.csv_btn = ctk.CTkButton(monthly_frame, text="Select CSV", command=self.select_csv_one)
        self.csv_btn.pack(pady=5, fill="x")

        self.gen_month_btn = ctk.CTkButton(monthly_frame, text="Generate Monthly Schedule", command=self.generate_monthly_schedule, fg_color="#4CAF50")
        self.gen_month_btn.pack(pady=10, fill="x")

        weekly_frame = ctk.CTkFrame(self.main_frame)
        weekly_frame.pack(pady=10, padx=10, fill="x")
        ctk.CTkLabel(weekly_frame, text="Modify Existing Weekly Schedule", font=("Helvetica", 14)).pack(pady=5)

        self.weekly_csv_btn = ctk.CTkButton(weekly_frame, text="Select CSV", command=self.select_csv_two)
        self.weekly_csv_btn.pack(pady=5, fill="x")

        self.weekly_file_btn = ctk.CTkButton(weekly_frame, text="Select Weekly Schedule XLSX", command=self.select_weekly_schedule)
        self.weekly_file_btn.pack(pady=5, fill="x")

        self.deleted_shifts_btn = ctk.CTkButton(weekly_frame, text="Select Deleted Shifts XLSX", command=self.select_deleted_shifts)
        self.deleted_shifts_btn.pack(pady=5, fill="x")

        self.modify_btn = ctk.CTkButton(weekly_frame, text="Apply Changes to Weekly Schedule", command=self.modify_weekly_schedule, fg_color="#4CAF50")
        self.modify_btn.pack(pady=10, fill="x")

        self.spinner_label = ctk.CTkLabel(self.main_frame, text="")
        self.spinner_label.pack(pady=5)

        self.status_label = ctk.CTkLabel(self.main_frame, text="")
        self.status_label.pack()

    def select_csv_one(self):
        self.input_csv_one = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if self.input_csv_one:
            self.csv_btn.configure(text=f"CSV: {os.path.basename(self.input_csv_one)}")

    def select_csv_two(self):
        self.input_csv_two = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if self.input_csv_two:
            self.weekly_csv_btn.configure(text=f"CSV: {os.path.basename(self.input_csv_two)}")

    def select_weekly_schedule(self):
        self.weekly_schedule_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.weekly_schedule_file:
            self.weekly_file_btn.configure(text=f"Weekly: {os.path.basename(self.weekly_schedule_file)}")

    def select_deleted_shifts(self):
        self.deleted_shifts_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.deleted_shifts_file:
            self.deleted_shifts_btn.configure(text=f"Shifts: {os.path.basename(self.deleted_shifts_file)}")

    def start_spinner(self, text):
        self.spinner_running = True
        self.spinner_label.configure(text=text)
        def animate():
            frames = ["⏳", "🕐", "🕑", "🕒", "🕓", "🕔", "🕕", "🕖", "🕗", "🕘", "🕙", "🕚"]
            i = 0
            while self.spinner_running:
                self.spinner_label.configure(text=f"{frames[i % len(frames)]} {text}")
                i += 1
                time.sleep(0.2)
        threading.Thread(target=animate, daemon=True).start()

    def stop_spinner(self):
        self.spinner_running = False
        self.spinner_label.configure(text="")

    def generate_monthly_schedule(self):
        if not self.input_csv_one:
            messagebox.showerror("Error", "Please select a CSV file.")
            return

        output_dir = filedialog.askdirectory(
            title="Select Folder to Save Weekly Excel Files"
        )
        if not output_dir:
            return

        self.gen_month_btn.configure(state="disabled")
        self.csv_btn.configure(state="disabled")
        self.weekly_csv_btn.configure(state="disabled")
        self.modify_btn.configure(state="disabled")
        self.weekly_file_btn.configure(state="disabled")
        self.deleted_shifts_btn.configure(state="disabled")
        self.start_spinner("Generating weekly schedules...")

        def run():
            try:
                generate_monthly_schedule_from_csv(self.input_csv_one,
                    self.input_json, output_dir)

                for week in range(1, 5):
                    rows = []
                    with open(os.path.join(output_dir, f"week_{week}.txt"), "r", encoding="utf-8") as f:
                        location, room = None, None
                        for line in f:
                            line = line.strip()
                            if line.startswith("Локація:"):
                                location = line.split("Локація:")[1].strip()
                            elif line.startswith("Кабінет:"):
                                room = line.split("Кабінет:")[1].strip()
                            elif line.startswith("(") and ")" in line:
                                try:
                                    part, doctor = line.split(") - ")
                                    day, shift = eval(part + ")")
                                    rows.append((week, location, room, day, shift, doctor if doctor != "None" else ""))
                                except:
                                    pass

                    wb = Workbook()
                    ws = wb.active
                    ws.title = f"Week {week}"
                    ws.append(["Week", "Location", "Room", "Day", "Shift", "Doctor"])
                    for row in rows:
                        ws.append(row)

                    output_file = os.path.join(output_dir, f"week_{week}.xlsx")
                    wb.save(output_file)

                messagebox.showinfo("Success", f"All weekly schedules saved to:\n{output_dir}")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                self.stop_spinner()
                self.gen_month_btn.configure(state="normal")
                self.csv_btn.configure(state="normal")
                self.weekly_csv_btn.configure(state="normal")
                self.modify_btn.configure(state="normal")
                self.weekly_file_btn.configure(state="normal")
                self.deleted_shifts_btn.configure(state="normal")

        threading.Thread(target=run).start()

    def modify_weekly_schedule(self):
        if not all([self.input_csv_two, self.input_json, self.weekly_schedule_file, self.deleted_shifts_file]):
            messagebox.showerror("Error", "Please select all required files (CSV, weekly schedule, deleted shifts).")
            return

        self.gen_month_btn.configure(state="disabled")
        self.csv_btn.configure(state="disabled")
        self.weekly_csv_btn.configure(state="disabled")
        self.modify_btn.configure(state="disabled")
        self.weekly_file_btn.configure(state="disabled")
        self.deleted_shifts_btn.configure(state="disabled")
        self.start_spinner("Modifying weekly schedule...")

        def run():
            try:
                # 1. Convert XLSX to TXT
                temp_txt = self.weekly_schedule_file.replace(".xlsx", "_temp.txt")
                wb = openpyxl.load_workbook(self.weekly_schedule_file)
                ws = wb.active

                with open(temp_txt, "w", encoding="utf-8") as f:
                    last_loc, last_room = None, None
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        _, location, room, day, shift, doctor = row

                        # New location block
                        if location != last_loc:
                            f.write(f"Локація: {location}\n")
                            f.write("=" * 40 + "\n")
                            last_loc = location
                            last_room = None  # reset room tracking

                        # New cabinet block
                        if room != last_room :
                            if last_room != None:
                                f.write("-" * 30 + "\n")
                            f.write(f"Кабінет: {room}\n")
                            f.write("-" * 30 + "\n")
                            last_room = room

                        # Write shift
                        f.write(f"({day}, {shift}) - {doctor if doctor else 'None'}\n")

                    f.write("-" * 30 + "\n")  # optional footer for last room

                deleted_shifts = defaultdict(set)

                wb_deleted = openpyxl.load_workbook(self.deleted_shifts_file)
                ws_deleted = wb_deleted.active

                for row in ws_deleted.iter_rows(min_row=2, values_only=True):  # skip header
                    doctor, shift_str = row
                    if doctor and shift_str:
                        try:
                            parts = shift_str.strip().split(".")
                            if len(parts) == 2:
                                day, shift = map(int, parts)
                                deleted_shifts[doctor.strip()].add((day, shift))
                        except Exception as e:
                            print(f"⚠️ Could not parse shift '{shift_str}' for doctor '{doctor}': {e}")

                print(deleted_shifts)

                change_weekly_schedule(
                    self.input_csv_two,
                    self.input_json,
                    temp_txt,
                    deleted_shifts
                )

                rows = []
                with open(temp_txt, "r", encoding="utf-8") as f:
                    location, room = None, None
                    for line in f:
                        line = line.strip()
                        if line.startswith("Локація:"):
                            location = line.split("Локація:")[1].strip()
                        elif line.startswith("Кабінет:"):
                            room = line.split("Кабінет:")[1].strip()
                        elif line.startswith("(") and ")" in line:
                            try:
                                part, doctor = line.split(") - ")
                                day, shift = eval(part + ")")
                                rows.append((None, location, room, day, shift, doctor if doctor != "None" else ""))
                            except:
                                pass

                wb_out = openpyxl.Workbook()
                ws_out = wb_out.active
                ws_out.title = "Weekly Schedule"
                ws_out.append(["Week", "Location", "Room", "Day", "Shift", "Doctor"])
                for row in rows:
                    ws_out.append(row)
                wb_out.save(self.weekly_schedule_file)

                os.remove(temp_txt)
                messagebox.showinfo("Success", f"Weekly schedule updated:\n{self.weekly_schedule_file}")
            except Exception as e:
                traceback_str = traceback.format_exc()
                print("❌ Exception occurred:\n", traceback_str)
                messagebox.showerror("Modification Failed", traceback_str)
            finally:
                self.stop_spinner()
                self.gen_month_btn.configure(state="normal")
                self.csv_btn.configure(state="normal")
                self.weekly_csv_btn.configure(state="normal")
                self.modify_btn.configure(state="normal")
                self.weekly_file_btn.configure(state="normal")
                self.deleted_shifts_btn.configure(state="normal")

        threading.Thread(target=run).start()

if __name__ == "__main__":
    app = ScheduleApp()
    app.mainloop()
